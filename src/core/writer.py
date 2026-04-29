"""Дозапись блоков и таблицы счетов на YW2PF, маркеры [XA:…], автофильтр."""

from __future__ import annotations

import re
import pandas as pd
from loguru import logger
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.config import (
    ACCOUNTS_BLOCK_NAME,
    BLOCK_GAP,
    BLOCK_MARKER_PREFIX,
    BLOCK_MARKER_SUFFIX,
    ACCOUNT_COMPUTED_HEADER,
    AUTOFILTER_EXACT_SHEETS,
    AUTO_WIDTH_TARGET_SHEETS,
    DATA_START_ROW,
    HEADER_ROW,
    TARGET_SHEET,
    YQJDATA_FILTER_FIELD,
    YQJDATA_FILTER_VALUE,
)

bold = Font(bold=True)
_SHEET_S_ONE_CHAR_PF_RE = re.compile(r"^S.PF$")


def _strip_previous_run(ws: Worksheet) -> None:
    """Удаляет все строки от пустой строки перед первым маркером ``[XA:`` вниз.

    ТЗ: идемпотентность — повторный запуск убирает хвост прошлой обработки
    (маркеры в колонке A).
    """
    first_marker_row: int | None = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        val = cell.value
        if isinstance(val, str) and val.startswith(BLOCK_MARKER_PREFIX):
            first_marker_row = cell.row
            break
    if first_marker_row is None:
        return
    # с первой пустой строки-зазора (строка перед маркером) до конца листа
    ws.delete_rows(first_marker_row - 1, ws.max_row - first_marker_row + 2)
    ws.auto_filter.ref = None


def _find_last_nonempty_row(ws: Worksheet) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(1, (ws.max_column or 0) + 1)
        ):
            return r
    return 0


def _apply_workbook_calc_flags(wb: Workbook) -> None:
    """Включает пересчёт при открытии в Excel (по плану §6.7)."""
    try:
        calc = wb.calculation
        if calc is not None:
            calc.fullCalcOnLoad = True  # type: ignore[assignment]
    except (AttributeError, TypeError) as e:
        logger.debug("Не удалось выставить fullCalcOnLoad: {}", e)
    try:
        props = wb.properties
        if props is not None and hasattr(props, "calcMode"):
            props.calcMode = "auto"  # type: ignore[assignment]
    except (AttributeError, TypeError) as e:
        logger.debug("Не удалось выставить calcMode: {}", e)


def apply_post_formatting(wb: Workbook, *, target_sheet: str) -> None:
    """Применяет пост-форматирование: автоширина + автофильтры по списку листов."""
    if target_sheet in AUTO_WIDTH_TARGET_SHEETS and target_sheet in wb.sheetnames:
        _autofit_sheet_columns(wb[target_sheet])
    _apply_autofilters_to_pf_sheets(wb)


def _autofit_sheet_columns(ws: Worksheet, *, min_width: int = 8, max_width: int = 60) -> None:
    """Подбирает ширину колонок по максимальной длине значений в строках 2..max_row."""
    last_col = 0
    for c in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
            last_col = c
    if last_col == 0:
        return

    for c in range(1, last_col + 1):
        max_len = 0
        for r in range(HEADER_ROW, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            ln = len(str(val).strip())
            if ln > max_len:
                max_len = ln
        width = min(max(min_width, max_len + 2), max_width)
        ws.column_dimensions[get_column_letter(c)].width = width


def _apply_autofilters_to_pf_sheets(wb: Workbook) -> None:
    """Ставит автофильтр на S?PF и на фиксированный список листов."""
    for name in wb.sheetnames:
        if not _sheet_needs_autofilter(name):
            continue
        _set_sheet_autofilter(wb[name])


def _sheet_needs_autofilter(sheet_name: str) -> bool:
    return bool(_SHEET_S_ONE_CHAR_PF_RE.match(sheet_name)) or sheet_name in AUTOFILTER_EXACT_SHEETS


def _set_sheet_autofilter(ws: Worksheet) -> None:
    """Устанавливает ref автофильтра от A2 до последней колонки/строки данных."""
    last_col = 0
    for c in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
            last_col = c
    if last_col == 0:
        return

    last_data_row = HEADER_ROW
    for r in range(ws.max_row, DATA_START_ROW - 1, -1):
        if any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(1, last_col + 1)
        ):
            last_data_row = r
            break
    end_col = get_column_letter(last_col)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{end_col}{last_data_row}"


def write_to_yw2pf(
    wb: Workbook,
    ordered_blocks: list[tuple[str, list, list[dict]]],
    account_df: pd.DataFrame | None,
    *,
    target_sheet: str | None = None,
    yqj_df: pd.DataFrame | None = None,
) -> None:
    """Пишет блоки листов и опционально таблицу счетов на целевой лист.

    ТЗ: не трогать существующие строки выше; маркер ``[XA:`` — граница
    идемпотентности; таблица счетов с формулой ``-(S5AIMD+S5AM1D)``;
    автофильтр на таблицу ``ACCOUNTS`` (YW-режим) или ``YQJDATA`` (YQ-режим).

    Args:
        wb: Книга (``data_only=False``).
        ordered_blocks: Список ``(имя_блока, headers, rows)`` — ``headers`` из строки 2
            исходного листа, ``rows`` — список dict по данным.
        account_df: Таблица из ``build_account_table``; если ``None`` или пустая —
            блок ``ACCOUNTS`` не пишется.
        target_sheet: Имя целевого листа; по умолчанию ``TARGET_SHEET`` (YW2PF).
        yqj_df: Объединённая таблица YQJDATA (только YQ-режим); если передана —
            пишется после ACCOUNTS с автофильтром и pre-set YQJSTS ≠ «А».
    """
    resolved_target_sheet = target_sheet or TARGET_SHEET
    ws = wb[resolved_target_sheet]
    _strip_previous_run(ws)

    cursor = _find_last_nonempty_row(ws) + 1 + BLOCK_GAP

    for block_name, headers, rows in ordered_blocks:
        ws.cell(
            row=cursor,
            column=1,
            value=f"{BLOCK_MARKER_PREFIX}{block_name}{BLOCK_MARKER_SUFFIX}",
        ).font = bold
        cursor += 1
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=cursor, column=ci, value=h).font = bold
        cursor += 1
        for row_dict in rows:
            for ci, h in enumerate(headers, start=1):
                ws.cell(row=cursor, column=ci, value=row_dict.get(h))
            cursor += 1
        cursor += BLOCK_GAP

    sheet_label = resolved_target_sheet
    if account_df is None or account_df.empty:
        logger.info("Счетов нет — блок ACCOUNTS на {} не записывается.", sheet_label)
    else:
        cursor = _write_account_table(ws, account_df, cursor)

    if yqj_df is not None and not yqj_df.empty:
        _write_yqjdata_block(ws, yqj_df, cursor)

    apply_post_formatting(wb, target_sheet=resolved_target_sheet)
    _apply_workbook_calc_flags(wb)


def _write_account_table(ws: Worksheet, account_df: pd.DataFrame, start_row: int) -> int:
    """Пишет ``[XA:ACCOUNTS]``, заголовки, данные, формулу, автофильтр."""
    cursor = start_row
    ws.cell(
        row=cursor,
        column=1,
        value=f"{BLOCK_MARKER_PREFIX}{ACCOUNTS_BLOCK_NAME}{BLOCK_MARKER_SUFFIX}",
    ).font = bold
    cursor += 1

    acc_headers = list(account_df.columns) + [ACCOUNT_COMPUTED_HEADER]
    header_row = cursor
    for ci, h in enumerate(acc_headers, start=1):
        ws.cell(row=cursor, column=ci, value=h).font = bold
    cursor += 1

    try:
        s5aimd_i = list(account_df.columns).index("S5AIMD") + 1
        s5am1d_i = list(account_df.columns).index("S5AM1D") + 1
    except ValueError as e:
        raise ValueError("В таблице счетов ожидаются колонки S5AIMD и S5AM1D.") from e
    s5aimd_letter = get_column_letter(s5aimd_i)
    s5am1d_letter = get_column_letter(s5am1d_i)
    neg_col_idx = len(account_df.columns) + 1

    for _, rec in account_df.iterrows():
        for ci, col in enumerate(account_df.columns, start=1):
            val = rec[col]
            out: object = None
            if not pd.isna(val):
                out = val
            ws.cell(row=cursor, column=ci, value=out)
        ws.cell(
            row=cursor,
            column=neg_col_idx,
            value=f"=-({s5aimd_letter}{cursor}+{s5am1d_letter}{cursor})",
        )
        cursor += 1
    data_end = cursor - 1
    if data_end < header_row:
        return cursor

    start_letter = get_column_letter(1)
    end_letter = get_column_letter(neg_col_idx)
    ws.auto_filter.ref = f"{start_letter}{header_row}:{end_letter}{data_end}"
    return cursor


def _write_yqjdata_block(ws: Worksheet, yqj_df: pd.DataFrame, start_row: int) -> int:
    """Пишет блок ``[XA:YQJDATA]`` с автофильтром и pre-set YQJSTS ≠ «А».

    ТЗ: «вывести данные YQJPF, YQJOPF, AN41PF; фильтр на всю структуру;
    фильтр YQJSTS ≠ «А»».

    Args:
        ws: Целевой лист (YQ2PF).
        yqj_df: DataFrame из ``build_yqj_table``.
        start_row: Строка, с которой начинается запись (включая BLOCK_GAP-отступ).

    Returns:
        Следующая свободная строка после последней строки данных.
    """
    from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn

    cursor = start_row + BLOCK_GAP
    ws.cell(
        row=cursor,
        column=1,
        value=f"{BLOCK_MARKER_PREFIX}YQJDATA{BLOCK_MARKER_SUFFIX}",
    ).font = bold
    cursor += 1

    headers = list(yqj_df.columns)
    header_row = cursor
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=cursor, column=ci, value=h).font = bold
    cursor += 1

    for _, rec in yqj_df.iterrows():
        for ci, col in enumerate(headers, start=1):
            val = rec[col]
            ws.cell(row=cursor, column=ci, value=val if pd.notna(val) else None)
        cursor += 1
    data_end = cursor - 1

    end_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{end_col_letter}{data_end}"

    if YQJDATA_FILTER_FIELD in headers:
        col_idx = headers.index(YQJDATA_FILTER_FIELD)  # 0-based для FilterColumn
        fc = FilterColumn(colId=col_idx)
        fc.customFilters = CustomFilters(
            customFilter=[CustomFilter(operator="notEqual", val=YQJDATA_FILTER_VALUE)]
        )
        ws.auto_filter.filterColumn.append(fc)

    return cursor
