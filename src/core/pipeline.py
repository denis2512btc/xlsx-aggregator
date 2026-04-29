"""Оркестрация: загрузка книги, валидация, JOIN, запись, безопасное сохранение."""

from __future__ import annotations

import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.core import joiner, writer
from src.core.account_extractor import extract_all_accounts, extract_all_accounts_yq
from src.core.config import (
    ALWAYS_APPEND,
    ALWAYS_APPEND_YQ,
    ACCOUNT_SOURCE_SHEETS,
    ACCOUNT_SOURCE_SHEETS_YQ,
    CONDITIONAL_APPEND,
    MAKE_BACKUP,
    S5_SHEET,
    SC_SHEET,
    TARGET_SHEET,
    TARGET_SHEET_YQ,
    YQJDATA_BASE_SHEET,
    YQJDATA_OPT_SHEET1,
    YQJDATA_OPT_SHEET2,
    YQ_CONDITIONAL_NUMERIC_SHEET,
    YQ_CONDITIONAL_STRING_SHEET,
    YQ_CONDITIONAL_TRIGGER,
)
from src.core.sheet_reader import read_sheet_as_dicts, sheet_headers_list

ProgressFn = Callable[[int, str], None]


def _is_blank(v: object) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == ""


def detect_mode(wb: Workbook) -> str:
    """Возвращает ``'YQ'`` если в книге есть лист YQ2PF, иначе ``'YW'``."""
    return "YQ" if TARGET_SHEET_YQ in wb.sheetnames else "YW"


def _is_numeric(v: object) -> bool:
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
        return True
    except (ValueError, TypeError):
        return False


def determine_yq_conditional_sheet(yq2pf_rows: list[dict]) -> str | None:
    """Возвращает имя условного листа для YQ-режима или ``None``.

    ТЗ: если ``YQ2PR2`` цифровое → AN4PF; символьное → AN6PF; пусто → None.
    """
    if not yq2pf_rows:
        return None
    v = yq2pf_rows[0].get(YQ_CONDITIONAL_TRIGGER)
    if _is_blank(v):
        return None
    return YQ_CONDITIONAL_NUMERIC_SHEET if _is_numeric(v) else YQ_CONDITIONAL_STRING_SHEET


def _run_yq_pipeline(wb: Workbook, path: Path, progress: "ProgressFn | None") -> "PipelineResult":
    """YQ-ветка пайплайна: целевой лист YQ2PF.

    ТЗ: YQ-режим активируется при наличии листа YQ2PF вместо YW2PF;
    источники счетов — YQ2PF, YQ3PF; условный лист — AN4PF или AN6PF;
    после ACCOUNTS — блок YQJDATA.
    """
    _notify(progress, 10, "YQ-режим: чтение листов")

    if TARGET_SHEET_YQ not in wb.sheetnames:
        raise RuntimeError(f"В книге отсутствует лист {TARGET_SHEET_YQ}.")
    if SC_SHEET not in wb.sheetnames:
        raise RuntimeError("В книге отсутствует обязательный лист SCPF.")

    yq2pf_rows = read_sheet_as_dicts(wb, TARGET_SHEET_YQ)
    cond_sheet = determine_yq_conditional_sheet(yq2pf_rows)

    sheets: dict[str, list[dict]] = {}
    for sn in ACCOUNT_SOURCE_SHEETS_YQ:
        if sn in wb.sheetnames:
            sheets[sn] = read_sheet_as_dicts(wb, sn)

    _notify(progress, 25, "YQ-режим: извлечение счетов")
    accounts = extract_all_accounts_yq(sheets)

    sc_rows = read_sheet_as_dicts(wb, SC_SHEET)
    s5_rows = read_sheet_as_dicts(wb, S5_SHEET) if S5_SHEET in wb.sheetnames else []
    if not s5_rows:
        logger.warning("Лист S5PF отсутствует или пуст — колонки S5* будут пустыми.")

    acc_df = joiner.build_account_table(
        accounts, sc_rows, s5_rows,
        col_order=joiner._ACCOUNT_COL_ORDER_YQ,
    ) if accounts else None
    _notify(progress, 40, f"Счетов: {len(accounts)}")

    ordered_blocks: list[tuple[str, list, list[dict]]] = []

    for sn in ALWAYS_APPEND_YQ:
        if sn in wb.sheetnames:
            ordered_blocks.append((sn, list(sheet_headers_list(wb, sn)), read_sheet_as_dicts(wb, sn)))
        else:
            logger.warning("Лист %s отсутствует — блок пропущен.", sn)

    if cond_sheet:
        if cond_sheet in wb.sheetnames:
            cond_rows = read_sheet_as_dicts(wb, cond_sheet)
            if cond_rows:
                ordered_blocks.append((cond_sheet, list(sheet_headers_list(wb, cond_sheet)), cond_rows))
            else:
                logger.info("Лист %s пустой — не выводится.", cond_sheet)
        else:
            logger.warning("Условный лист %s отсутствует.", cond_sheet)

    _notify(progress, 55, "YQ-режим: построение YQJDATA")
    yqj_df = None
    if YQJDATA_BASE_SHEET in wb.sheetnames:
        yqjpf_rows = read_sheet_as_dicts(wb, YQJDATA_BASE_SHEET)
        if yqjpf_rows:
            yqjopf_rows = (
                read_sheet_as_dicts(wb, YQJDATA_OPT_SHEET1)
                if YQJDATA_OPT_SHEET1 in wb.sheetnames else []
            )
            an41pf_rows = (
                read_sheet_as_dicts(wb, YQJDATA_OPT_SHEET2)
                if YQJDATA_OPT_SHEET2 in wb.sheetnames else []
            )
            yqj_df = joiner.build_yqj_table(yqjpf_rows, yqjopf_rows, an41pf_rows)
        else:
            logger.warning("YQJPF пустой — блок YQJDATA не пишется.")
    else:
        logger.warning("Лист %s отсутствует — блок YQJDATA не пишется.", YQJDATA_BASE_SHEET)

    _notify(progress, 70, f"Запись в {TARGET_SHEET_YQ}")
    writer.write_to_yw2pf(wb, ordered_blocks, acc_df,
                          target_sheet=TARGET_SHEET_YQ, yqj_df=yqj_df)

    _notify(progress, 85, "Сохранение…")
    meta = _safe_overwrite_save(wb, str(path))
    _notify(progress, 100, f"Готово. Бэкап: {meta['backup']}")
    return PipelineResult(
        result_path=meta["result"],
        backup_path=meta["backup"],
        account_count=len(accounts),
    )


def _notify(progress: ProgressFn | None, pct: int, msg: str) -> None:
    if progress:
        progress(pct, msg)
    logger.info(msg)


def _assert_file_writable(path: Path) -> None:
    """Проверка блокировки файла (например, открыт в Excel) — до бэкапа и изменений."""
    try:
        with open(path, "ab"):
            pass
    except PermissionError as e:
        raise RuntimeError(
            f"Файл занят другим процессом (вероятно открыт в Excel): {path}. "
            "Закройте файл и повторите."
        ) from e


def _read_yw2pf_first_data_row_triggers(
    xlsx_path: Path, trigger_fields: set[str]
) -> dict[str, object]:
    """Читает значения триггеров из первой data-строки YW2PF (формулы — кэш data_only)."""
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        if TARGET_SHEET not in wb.sheetnames:
            return {}
        ws = wb[TARGET_SHEET]
        headers = [c.value for c in ws[2]]
        if not headers:
            return {}
        hmap = {h: i for i, h in enumerate(headers) if h is not None}
        row3 = [ws.cell(row=3, column=i + 1).value for i in range(len(headers))]
        out: dict[str, object] = {}
        for f in trigger_fields:
            if f in hmap:
                idx = hmap[f]
                if idx < len(row3):
                    out[f] = row3[idx]
        return out
    finally:
        wb.close()


def _headers_for_block(wb: Workbook, sheet_name: str) -> list:
    if sheet_name not in wb.sheetnames:
        return []
    return list(sheet_headers_list(wb, sheet_name))


def _safe_overwrite_save(wb: Workbook, original_path: str) -> dict[str, str]:
    """Сохраняет ``wb`` поверх ``original_path`` с бэкапом и атомарной подменой.

    Returns:
        ``{'result': путь_к_файлу, 'backup': путь_к_бэкапу}``
    """
    orig = Path(original_path)
    _assert_file_writable(orig)

    if MAKE_BACKUP:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = orig.with_name(f"{orig.stem}.backup_{ts}{orig.suffix}")
        shutil.copy2(orig, backup)
    else:
        backup = orig

    with tempfile.NamedTemporaryFile(
        dir=orig.parent,
        prefix=f".{orig.stem}.",
        suffix=".tmp",
        delete=False,
    ) as tf:
        tmp_path = Path(tf.name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, orig)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink()
        raise
    if MAKE_BACKUP:
        return {"result": str(orig), "backup": str(backup)}
    return {"result": str(orig), "backup": str(backup)}


@dataclass
class PipelineResult:
    """Результат успешного прогона пайплайна."""

    result_path: str
    backup_path: str
    account_count: int


def run_pipeline(
    xlsx_path: str | Path,
    progress: ProgressFn | None = None,
) -> PipelineResult:
    """Выполняет полный цикл обработки и перезаписи файла.

    ТЗ: см. [PLAN.md] разд. 6–7 — загрузка, извлечение счетов, JOIN, запись,
    бэкап, атомарная замена.

    Args:
        xlsx_path: Путь к ``.xlsx`` (будет перезаписан).
        progress: Необязательный callback ``(процент, сообщение)``.

    Returns:
        ``PipelineResult`` с путями к файлу и бэкапу.

    Raises:
        RuntimeError: Файл занят, нет обязательных листов и т.п.
        OSError: Ошибка чтения/записи.
    """
    path = Path(xlsx_path).resolve()
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Ожидается файл .xlsx")

    _notify(progress, 0, "Старт обработки")
    _assert_file_writable(path)

    trigger_field_names = {c["trigger_field"] for c in CONDITIONAL_APPEND}
    triggers = _read_yw2pf_first_data_row_triggers(path, trigger_field_names)

    wb = load_workbook(path, data_only=False)
    try:
        mode = detect_mode(wb)
        logger.info("Режим обработки: %s", mode)
        if mode == "YQ":
            return _run_yq_pipeline(wb, path, progress)

        if TARGET_SHEET not in wb.sheetnames:
            raise RuntimeError("В книге отсутствует обязательный лист YW2PF.")
        if SC_SHEET not in wb.sheetnames:
            raise RuntimeError("В книге отсутствует обязательный лист SCPF.")

        have_s5 = S5_SHEET in wb.sheetnames
        if not have_s5:
            logger.warning("Лист S5PF отсутствует — колонки S5* будут пустыми (штатно по ТЗ).")

        _notify(progress, 10, "Книга загружена")

        sc_rows = read_sheet_as_dicts(wb, SC_SHEET) if SC_SHEET in wb.sheetnames else []
        s5_rows = read_sheet_as_dicts(wb, S5_SHEET) if have_s5 else []

        # Данные для извлечения счетов: только листы-источники
        sheet_data: dict[str, list[dict]] = {}
        for sn in ACCOUNT_SOURCE_SHEETS:
            if sn not in wb.sheetnames:
                raise RuntimeError(f"Отсутствует лист, необходимый для счетов: {sn}")
            sheet_data[sn] = read_sheet_as_dicts(wb, sn)

        accounts = extract_all_accounts(sheet_data)
        _notify(progress, 30, f"Извлечено уникальных счетов: {len(accounts)}")

        acc_df: pd.DataFrame | None
        if not accounts:
            acc_df = None
        else:
            acc_df = joiner.build_account_table(accounts, sc_rows, s5_rows)
        _notify(progress, 50, "Таблица счетов построена")

        def _trigger_value(field: str) -> object:
            v = triggers.get(field)
            if v is None and sheet_data.get(TARGET_SHEET) and sheet_data[TARGET_SHEET]:
                v = sheet_data[TARGET_SHEET][0].get(field)
            return v

        conditional_include = {
            item["sheet"]: not _is_blank(_trigger_value(item["trigger_field"]))
            for item in CONDITIONAL_APPEND
        }

        ordered_blocks: list[tuple[str, list, list[dict]]] = []

        for ap_sheet in ALWAYS_APPEND:
            if ap_sheet not in wb.sheetnames:
                logger.warning("Лист %s отсутствует — блок пропущен.", ap_sheet)
                continue
            headers = _headers_for_block(wb, ap_sheet)
            rows = read_sheet_as_dicts(wb, ap_sheet)
            ordered_blocks.append((ap_sheet, headers, rows))

        for item in CONDITIONAL_APPEND:
            sh = item["sheet"]
            if not conditional_include.get(sh, False):
                logger.info("Условие для листа %s ложно — блок пропущен.", sh)
                continue
            if sh not in wb.sheetnames:
                logger.warning("Условный лист %s отсутствует — блок пропущен.", sh)
                continue
            headers = _headers_for_block(wb, sh)
            rows = read_sheet_as_dicts(wb, sh)
            ordered_blocks.append((sh, headers, rows))

        _notify(progress, 70, "Запись на YW2PF")
        writer.write_to_yw2pf(wb, ordered_blocks, acc_df)
        _notify(progress, 90, "Сохранение…")

        meta = _safe_overwrite_save(wb, str(path))
        _notify(progress, 100, "Готово")
        return PipelineResult(
            result_path=meta["result"],
            backup_path=meta["backup"],
            account_count=len(accounts),
        )
    finally:
        wb.close()


def main_cli() -> None:
    """Точка входа для ``python -m src.core.pipeline <файл.xlsx>``."""
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m src.core.pipeline <file.xlsx>", file=sys.stderr)
        raise SystemExit(1)
    p = Path(sys.argv[1])
    r = run_pipeline(p, progress=lambda n, t: print(f"{n}%: {t}"))
    print("OK", r)
    


if __name__ == "__main__":
    main_cli()
