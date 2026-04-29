"""Интеграционный тест YQ-режима на реальном файле AMSAV."""

from __future__ import annotations

import shutil

import pytest
from openpyxl import load_workbook

from src.core.config import (
    AUTOFILTER_EXACT_SHEETS,
    BLOCK_MARKER_PREFIX,
    TARGET_SHEET_YQ,
)
from src.core.pipeline import detect_mode, run_pipeline

FIXTURE = "tests/fixtures/AMSAV2026-04-16-16.28.32.264864.xlsx"


@pytest.fixture
def yq_file(tmp_path):
    dst = tmp_path / "test_yq.xlsx"
    shutil.copy(FIXTURE, dst)
    return str(dst)


def test_detect_mode_yq():
    wb = load_workbook(FIXTURE, read_only=True, data_only=True)
    try:
        assert detect_mode(wb) == "YQ"
    finally:
        wb.close()


def test_yq_pipeline_produces_all_blocks(yq_file):
    result = run_pipeline(yq_file)

    assert result.result_path == yq_file
    assert result.backup_path == yq_file

    wb = load_workbook(yq_file)
    try:
        assert TARGET_SHEET_YQ in wb.sheetnames, "Лист YQ2PF отсутствует"
        ws = wb[TARGET_SHEET_YQ]

        markers = [
            ws.cell(row=r, column=1).value
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=1).value, str)
            and ws.cell(row=r, column=1).value.startswith(BLOCK_MARKER_PREFIX)
        ]

        assert any("YQ3PF" in m for m in markers), f"Блок YQ3PF не найден; маркеры: {markers}"
        assert any("ACCOUNTS" in m for m in markers), f"Блок ACCOUNTS не найден; маркеры: {markers}"
        assert any("YQJDATA" in m for m in markers), f"Блок YQJDATA не найден; маркеры: {markers}"

        assert ws.auto_filter.ref is not None, "Автофильтр не наложен"
        assert ws.column_dimensions["A"].width is not None, "Ширина колонок YQ2PF не выставлена"
        assert wb.active.title == TARGET_SHEET_YQ, "Активный лист должен быть YQ2PF"

        required_filter_sheets = {"SCPF", "S5PF", "YYR6PF", "YSAPF", "YR7PF", "YQJPF", "YQKPF", "JUHPF", "JUPF"}
        for name in required_filter_sheets:
            assert name in wb.sheetnames, f"В фикстуре отсутствует лист {name}"
            assert wb[name].auto_filter.ref is not None, f"Автофильтр не установлен на {name}"

        for name in AUTOFILTER_EXACT_SHEETS:
            if name in wb.sheetnames:
                assert wb[name].auto_filter.ref is not None, f"Автофильтр не установлен на {name}"
    finally:
        wb.close()


def test_yq_pipeline_rejects_already_processed_file(yq_file):
    wb = load_workbook(yq_file)
    try:
        ws = wb[TARGET_SHEET_YQ]
        ws.cell(row=4, column=1, value="already-processed-tail")
        wb.save(yq_file)
    finally:
        wb.close()

    with pytest.raises(RuntimeError, match="Файл уже обрабатывался"):
        run_pipeline(yq_file)


def test_yq_pipeline_second_run_blocked(yq_file):
    """Повторный запуск блокируется по правилу «файл уже обрабатывался»."""
    run_pipeline(yq_file)
    with pytest.raises(RuntimeError, match="Файл уже обрабатывался"):
        run_pipeline(yq_file)
